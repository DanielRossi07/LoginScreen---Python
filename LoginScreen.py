import tkinter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import openpyxl

# Designing window for login
def login_screen_function():
    global main_screen
    main_screen.withdraw()

    global login_screen
    global username_entry
    global password_entry

    login_screen = tkinter.Toplevel()
    login_screen.geometry('300x200')
    login_screen.title('Login Screen')

    welcome_label = tkinter.Label(login_screen, text='Welcome to Rossix')
    welcome_label.place(x=150, y=30, anchor='center')

    username_label = tkinter.Label(login_screen, text='Username: ')
    username_label.place(x=10, y=50)
    username_entry = tkinter.StringVar()
    username_entry = tkinter.Entry(login_screen, width=25, textvariable=username_entry)
    username_entry.place(x=90, y=52)

    password_label = tkinter.Label(login_screen, text='Password: ')
    password_label.place(x=10, y=90)
    password_entry = tkinter.StringVar()
    password_entry = tkinter.Entry(login_screen, width=25, textvariable=password_entry, show='*')
    password_entry.place(x=90, y=92)

    login_button = tkinter.Button(login_screen, text='Login', width=7, command=login_function)
    login_button.place(x=90, y=150, anchor='center')

    return_button = tkinter.Button(login_screen, text='Return', width=7, command=return_login_function)
    return_button.place(x=215, y=150, anchor='center')

# Designing window for registration
def register_screen_function():
    global main_screen
    main_screen.withdraw()

    global register_screen
    global username_entry
    global password_entry
    global confirmation_entry

    register_screen = tkinter.Toplevel()
    register_screen.geometry('300x240')
    register_screen.title('Register Screen')

    welcome_label = tkinter.Label(register_screen, text='Welcome to Rossix')
    welcome_label.place(x=150, y=30, anchor='center')

    username_label = tkinter.Label(register_screen, text='Username: ')
    username_label.place(x=10, y=50)
    username_entry = tkinter.StringVar()
    username_entry = tkinter.Entry(register_screen, width=25, textvariable=username_entry)
    username_entry.place(x=100, y=52)

    password_label = tkinter.Label(register_screen, text='Password: ')
    password_label.place(x=10, y=90)
    password_entry = tkinter.StringVar()
    password_entry = tkinter.Entry(register_screen, width=25, textvariable=password_entry, show='*')
    password_entry.place(x=100, y=92)

    confirmation_label = tkinter.Label(register_screen, text='Confirmation: ')
    confirmation_label.place(x=10, y=130)
    confirmation_entry = tkinter.StringVar()
    confirmation_entry = tkinter.Entry(register_screen, width=25, textvariable=confirmation_entry, show='*')
    confirmation_entry.place(x=100, y=132)

    register_button = tkinter.Button(register_screen, text='Register', width=7, command=register_function)
    register_button.place(x=90, y=190, anchor='center')

    return_button = tkinter.Button(register_screen, text='Return', width=7, command=return_register_function)
    return_button.place(x=215, y=190, anchor='center')


# Designing window for after login
def main_program_screen_function():
    global main_program_screen

    main_program_screen = tkinter.Toplevel()
    main_program_screen.geometry('300x300')
    main_program_screen.title('Main Program')

    label = tkinter.Label(main_program_screen, text='Welcome to ...')
    label.place(x=150, y=100, anchor='center')

    button = tkinter.Button(main_program_screen, text='Close Program', command=close_main_program_function)
    button.place(x=150, y=200, anchor='center')


# Implementing event on register button
def register_function():
    global username_entry
    global password_entry
    global confirmation_entry

    username = username_entry.get().lower()
    password = password_entry.get()
    confirmation = confirmation_entry.get()

    if len(username) == 0:
        error_blank_field_function()
    elif len(password) == 0:
        error_blank_field_function()
    elif len(confirmation) == 0:
        error_blank_field_function()
    else:
        if password != confirmation:
            error_password_match_function()
        else:
            try:
                # Create workbook instance
                wb = Workbook()
                # Load existing workbook
                wb = load_workbook('LoginData.xlsx')
                # Create active worksheet
                ws = wb.active
                pos = str(ws.max_row + 1)
            except:
                wb = Workbook()
                wb.save('LoginData.xlsx')
                ws = wb.active
                pos = str(ws.max_row)
            ws['A' + pos] = username
            ws['B' + pos] = password
            wb.save('LoginData.xlsx')
            register_success_screen_function()


# Implementing event on Login button
def login_function():
    #global main_screen

    global username_entry
    global password_entry

    username = username_entry.get().lower()
    password = password_entry.get()

    count = 0

    if len(username) == 0:
        error_blank_field_function()
    elif len(password) == 0:
        error_blank_field_function()
    else:
        username_validator()
        if username_validator() == True:
            password_validator()
            if password_validator() == True:
                login_success_screen_function()
            else:
                wrong_password_function()
        else:
            username_not_found_function()


# Function to validade password
def password_validator():
    global password_entry

    password = password_entry.get()

    wb = Workbook()
    wb = load_workbook('LoginData.xlsx')
    ws = wb.active
    max = ws.max_row + 1
    for cell in range(1, max):
        pos = str(cell)
        if ws['B' + pos].value == password:
            return True
    return False


# Function to validate username
def username_validator():
    global username_entry

    username = username_entry.get().lower()
    try:
        wb = Workbook()
        wb = load_workbook('LoginData.xlsx')
        ws = wb.active
        max = ws.max_row + 1
        for cell in range(1, max):
            pos = str(cell)
            if ws['A' + pos].value == username:
                return True
        return False
    except:
        return False


# Function to return from login
def return_login_function():
    login_screen.destroy()
    main_screen_function()

# Function to return from register
def return_register_function():
    register_screen.destroy()
    main_screen_function()


# Designing popup for successful login
def login_success_screen_function():
    global login_success_screen

    login_success_screen = tkinter.Toplevel()
    login_success_screen.geometry('300x150')
    login_success_screen.title('Login message!')

    label = tkinter.Label(login_success_screen, text='Successfully logged in!')
    label.place(x=150, y=30, anchor='center')

    button = tkinter.Button(login_success_screen, text='Ok', command=destroy_login_success_screen_function)
    button.place(x=150, y=80, anchor='center')


# Designing popup for successful register
def register_success_screen_function():
    global register_success_screen

    register_success_screen = tkinter.Toplevel()
    register_success_screen.geometry('300x150')
    register_success_screen.title('Register message')

    label = tkinter.Label(register_success_screen, text='Account successfully registered!')
    label.place(x=150, y=30, anchor='center')

    button = tkinter.Button(register_success_screen, text='Ok', command=destroy_register_success_screen_function)
    button.place(x=150, y=80, anchor='center')


# Designing popup for username not found
def username_not_found_function():
    global username_not_found

    username_not_found = tkinter.Toplevel()
    username_not_found.geometry('300x150')
    username_not_found.title('Error message')

    erro_label = tkinter.Label(username_not_found, text='User not found')
    erro_label.place(x=150, y=30, anchor='center')

    error_button = tkinter.Button(username_not_found, text='Ok', command=destroy_username_not_found_function)
    error_button.place(x=150, y=80, anchor='center')


# Designing popup for wrong password
def wrong_password_function():
    global error_password

    error_password = tkinter.Toplevel()
    error_password.geometry('300x150')
    error_password.title('Error')

    erro_label = tkinter.Label(error_password, text='Wrong Password')
    erro_label.place(x=150, y=30, anchor='center')

    error_button = tkinter.Button(error_password, text='Ok', command=destroy_error_password_function)
    error_button.place(x=150, y=80, anchor='center')


# Designing popup for wrong password and confirmation
def error_password_match_function():
    global error_password_match

    error_password_match = tkinter.Toplevel()
    error_password_match.geometry('300x150')
    error_password_match.title('Error')

    erro_label = tkinter.Label(error_password_match, text='Password and confirmation DOES NOT match!')
    erro_label.place(x=150, y=30, anchor='center')

    error_button = tkinter.Button(error_password_match, text='Ok', command=destroy_error_password_match_function)
    error_button.place(x=150, y=80, anchor='center')


# Designing popup for blank fields
def error_blank_field_function():
    global error_blank_field

    error_blank_field = tkinter.Toplevel()
    error_blank_field.geometry('300x150')
    error_blank_field.title('Error')

    erro_label = tkinter.Label(error_blank_field, text='Fill in all the fields, please.')
    erro_label.place(x=150, y=30, anchor='center')

    error_button = tkinter.Button(error_blank_field, text='Ok', command=destroy_error_blank_field_function)
    error_button.place(x=150, y=80, anchor='center')


# Deleting popups
def destroy_username_not_found_function():
    username_not_found.destroy()


def destroy_error_password_function():
    error_password.destroy()


def destroy_login_success_screen_function():
    login_success_screen.destroy()
    login_screen.destroy()
    main_program_screen_function()


def destroy_register_success_screen_function():
    register_success_screen.destroy()
    register_screen.destroy()
    main_screen.deiconify()


def destroy_error_password_match_function():
    error_password_match.destroy()


def destroy_error_blank_field_function():
    error_blank_field.destroy()


#Closing Main program
def close_main_program_function():
    main_program_screen.destroy()
    quit()

# Designing window for main (first) screen
def main_screen_function():
    global main_screen

    main_screen = tkinter.Tk()
    main_screen.geometry('200x100')
    main_screen.title('Main Screen')

    welcome_label = tkinter.Label(main_screen, text='Main Screen')
    welcome_label.place(x=100, y=30, anchor='center')

    login_button = tkinter.Button(main_screen, text='Login', width=7, command=login_screen_function)
    login_button.place(x=60, y=65, anchor='center')

    register_button = tkinter.Button(main_screen, text='Register', width=7, command=register_screen_function)
    register_button.place(x=140, y=65, anchor='center')

    main_screen.mainloop()


main_screen_function()